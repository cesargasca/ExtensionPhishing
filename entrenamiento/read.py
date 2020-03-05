# -*- coding: utf-8 -*-
"""
Created on Mon Mar  2 09:11:06 2020

@author: cgasca
"""
import csv
import numpy
import pandas as pd


def readCSV(filename):
    raw_data = open(filename, 'rt')
    reader = csv.reader(raw_data, delimiter=',', quoting=csv.QUOTE_NONE)
    x = list(reader)
    return x

def readXLSX(filename):
    df = pd.read_excel (filename,encoding='utf-8')
    df = df.values.tolist()
    y = []
    X = []
    for i in df:
        y.append(i[1])
        X.append(i[0])
    return X,y
    
def replaceNonAscii(string):
    return "".join(i for i in string if ord(i)<128)

    
def writeXLSX(filename,new_row):
    import openpyxl
    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.get_sheet_by_name('Hoja1')
    row = ws.max_row + 1
    
    for entry in new_row:
        if len(entry) < 12:
            print("--------------")
            print(entry)
        else:
            ws.cell(row=row, column = 1, value=entry)
            ws.cell(row=row, column = 2, value="phishing")
            row += 1
    
    wb.save(filename)
        
        
def readXML(filename):
    import quopri
    from lxml import etree
    with open(filename,encoding='utf-8') as f:
        root = etree.parse(f)
    emails = root.getroot().getchildren()
    text = []
    for e in emails:
        aux = replaceNonAscii(e.text)
        aux = quopri.decodestring(aux)
        aux = aux.decode('latin-1')
        text.append(aux)
    return text

if __name__ == "__main__":
    print("read")
    #X,y = readXLSX('Libro1.xlsx')
    text = readXML('spam.xml')
    writeXLSX('corpus.xlsx',text)


    